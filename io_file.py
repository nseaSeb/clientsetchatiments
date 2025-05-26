import csv
import os
import re
import openpyxl
from PySide6 import QtCore, QtWidgets, QtGui
from PySide6.QtCore import Qt, QTimer, Signal
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QTableWidget,
                              QTableWidgetItem, QMenuBar, QMenu, QFileDialog, QMessageBox,
                              QInputDialog, QLabel, QPushButton, QDialog, QRadioButton,QDockWidget,
                              QButtonGroup, QLineEdit, QCheckBox, QFrame,QTextEdit,QDialogButtonBox)
def open_file(self):
    def guess_separator(line):
        """Devine le séparateur le plus probable dans une ligne CSV"""
        candidates = [',', ';', '\t', '|']
        counts = {sep: line.count(sep) for sep in candidates}
        best_sep = max(counts, key=counts.get)
        return best_sep if counts[best_sep] > 0 else ';'

    def get_file_preview(file_path, encoding, sep, max_lines=3):
        """Retourne un aperçu du fichier avec l'encodage spécifié"""
        try:
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                preview = []
                reader = csv.reader(f, delimiter=sep)
                for i, row in enumerate(reader):
                    if i >= max_lines:
                        break
                    preview.append(row)
                return preview
        except:
            return None

    def show_encoding_dialog(file_path, sep):
        """Affiche une boîte de dialogue pour choisir l'encodage"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Sélection de l'encodage")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout()
        
        # Liste des encodages courants
        encodings = [
            ('UTF-8', 'utf-8'),
            ('Latin-1 (ISO-8859-1)', 'latin-1'),
            ('Windows-1252', 'windows-1252'),
            ('UTF-16', 'utf-16'),
            ('cp850', 'cp850'),
            ('cp1252', 'cp1252'),
            ('latin_1', 'latin_1'),
            ('mac_roman', 'mac_roman'),
            ('ascii', 'ascii'),
            ('Autre...', None)
        ]
        
        # Widgets
        encoding_group = QButtonGroup()
        preview = QTextEdit()
        preview.setReadOnly(True)
        preview.setLineWrapMode(QTextEdit.NoWrap)
        custom_encoding = QLineEdit()
        custom_encoding.setPlaceholderText("Entrez un encodage personnalisé")
        custom_encoding.setEnabled(False)
        
        # Ajout des options
        for i, (name, enc) in enumerate(encodings):
            radio = QRadioButton(name)
            encoding_group.addButton(radio, i)
            layout.addWidget(radio)
            
            # Aperçu pour cet encodage
            if enc:
                preview_text = get_file_preview(file_path, enc, sep)
                if preview_text:
                    radio.setToolTip("\n".join(" | ".join(cell for cell in row) for row in preview_text))
        
        # Connexion pour l'option "Autre"
        encoding_group.button(4).toggled.connect(custom_encoding.setEnabled)
        
        # Ajout des autres widgets
        layout.addWidget(QLabel("Aperçu des premières lignes:"))
        layout.addWidget(preview)
        layout.addWidget(custom_encoding)
        
        # Boutons OK/Annuler
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        # Mise à jour de l'aperçu quand on change d'option
        def update_preview():
            idx = encoding_group.checkedId()
            if idx == 4:  # Autre
                enc = custom_encoding.text().strip() or 'utf-8'
            else:
                enc = encodings[idx][1]
            
            preview_text = get_file_preview(file_path, enc, sep)
            if preview_text:
                preview.setPlainText("\n".join(" | ".join(str(cell) for cell in row) for row in preview_text))
            else:
                preview.setPlainText(f"ERREUR: Impossible de lire avec l'encodage {enc}")
        
        encoding_group.buttonClicked.connect(update_preview)
        encoding_group.buttons()[0].setChecked(True)
        update_preview()
        
        dialog.setLayout(layout)
        
        if dialog.exec() == QDialog.Accepted:
            idx = encoding_group.checkedId()
            if idx == 4:  # Autre
                return custom_encoding.text().strip() or 'utf-8'
            return encodings[idx][1]
        return None

    # Début de la fonction open_file
    self.table.setSortingEnabled(False)
    file_path, _ = QFileDialog.getOpenFileName(
        self,
        "Ouvrir un fichier Excel ou CSV",
        "",
        "Fichiers Excel ou CSV (*.xlsx *.csv);;Tous les fichiers (*.*)"
    )

    if not file_path:
        return

    try:
        ext = os.path.splitext(file_path)[1].lower()
        data = []

        if ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
        elif ext == ".csv":
            # Détection initiale du séparateur
            with open(file_path, 'rb') as f:
                first_line = f.readline().decode('latin-1')  # Lecture permissive
            sep = guess_separator(first_line)
            
            # Boîte de dialogue pour choisir l'encodage
            encoding = show_encoding_dialog(file_path, sep)
            if not encoding:  # Annulation
                return
                
            # Lecture finale avec l'encodage choisi
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f, delimiter=sep)
                data = [row for row in reader]
            
            self.show_message(f"Fichier CSV chargé (Encodage: {encoding}, Séparateur: '{sep}')")
        else:
            QMessageBox.critical(self, "Erreur", "Format de fichier non pris en charge.")
            return

        # Reste du traitement (identique)
        if data:
            self.setWindowTitle(f"{file_path} - {len(data)} lignes")
            self.headers = data[0]
            self.table.setRowCount(len(data) - 1)
            self.table.setColumnCount(len(self.headers))

            for col_idx, header in enumerate(self.headers):
                self.table.setHorizontalHeaderItem(col_idx, QTableWidgetItem(str(header)))

            for row_idx, row in enumerate(data[1:]):
                for col_idx, value in enumerate(row):
                    display_value = "" if value is None else str(value)
                    item = QTableWidgetItem(display_value)
                    self.table.setItem(row_idx, col_idx, item)

            self.table.setSortingEnabled(True)
            self.show_message(f"Fichier chargé : {file_path}")

    except Exception as e:
        QMessageBox.critical(self, "Erreur", f"Impossible de charger le fichier :\n{str(e)}")
# def open_file(self):
#         def guess_separator(line):
#             """Devine le séparateur le plus probable dans une ligne CSV"""
#             if not line:
#                 return ';' 
#             candidates = [',', ';', '\t', '|']
#             counts = {sep: line.count(sep) for sep in candidates}
#             best_sep = max(counts, key=counts.get)

#             # Si tous les counts sont à 0 (aucun séparateur trouvé)
#             if counts[best_sep] == 0:
#                 return ';'  # valeur par défaut

#             return best_sep
#         self.table.setSortingEnabled(False)
#         file_path, _ = QFileDialog.getOpenFileName(
#             self,
#             "Ouvrir un fichier Excel ou CSV",
#             "",
#             "Fichiers Excel ou CSV (*.xlsx *.csv);;Tous les fichiers (*.*)"
#         )

#         if not file_path:
#             return

#         try:
#             ext = os.path.splitext(file_path)[1].lower()
#             data = []

#             if ext == ".xlsx":
#                 wb = openpyxl.load_workbook(file_path)
#                 ws = wb.active
#                 for row in ws.iter_rows(values_only=True):
#                     data.append(list(row))
#             elif ext == ".csv":
#                 with open(file_path, newline='', encoding='utf-8') as f:
#                     first_line = f.readline()
#                     sep = guess_separator(first_line)
#                     self.show_message(f"Fichier CSV chargé avec séparateur détecté : '{sep}'")
#                     f.seek(0)  # Revenir au début
#                     reader = csv.reader(f, delimiter=sep)
#                     data = [row for row in reader]
#             else:
#                 QMessageBox.critical(self, "Erreur", "Format de fichier non pris en charge.")
#                 return

#             self.setWindowTitle(f"{file_path} - {len(data)} lignes")
#             self.headers = data[0]

#             self.table.setRowCount(len(data) - 1)
#             self.table.setColumnCount(len(self.headers))

#             for col_idx, header in enumerate(self.headers):
#                 self.table.setHorizontalHeaderItem(col_idx, QTableWidgetItem(str(header)))

#             for row_idx, row in enumerate(data[1:]):
#                 print(row_idx)
#                 for col_idx, value in enumerate(row):
#                     display_value = "" if value is None else str(value)
#                     item = QTableWidgetItem(str(display_value))
#                     self.table.setItem(row_idx, col_idx, item)
#             # Réactiver le tri une fois les données chargées
#             self.table.setSortingEnabled(True)

#             # Configurer le header pour permettre le tri
#             header = self.table.horizontalHeader()
#             header.setSortIndicatorShown(True)
#             header.setSectionsClickable(True)
#             self.show_message(f"Fichier chargé : {file_path}")
#             self.show_message(f"Nombre de lignes : {len(data) - 1}")

#         except Exception as e:
#             QMessageBox.critical(self, "Erreur", f"Impossible de charger le fichier :\n{str(e)}")

def save_file(self):
        if not self.has_data():
            QMessageBox.critical(self, "Erreur", "Aucune donnée à sauvegarder.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer sous",
            "",
            "Fichier CSV (*.csv);;Fichier Excel (*.xlsx)"
        )

        if not file_path:
            return

        ext = os.path.splitext(file_path)[1].lower()
        data = self.get_table_data()

        try:
            if ext == ".xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(self.headers)
                for row in data:
                    ws.append(row)
                wb.save(file_path)
            elif ext == ".csv":
                with open(file_path, mode="w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f, delimiter=';')
                    writer.writerow(self.headers)
                    writer.writerows(data)
            else:
                QMessageBox.critical(self, "Erreur", "Extension non prise en charge.")
                return

            self.show_message(f"Fichier sauvegardé avec succès : {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la sauvegarde : {str(e)}")

def save_segments(self):
        if not self.has_data():
            QMessageBox.critical(self, "Erreur", "Aucune donnée à sauvegarder.")
            return

        data = self.get_table_data()

        segment_size, ok = QInputDialog.getInt(
            self,
            "Taille segment",
            "Nombre de lignes par segment (hors en-tête) :",
            min=1
        )

        if not ok:
            return

        folder = QFileDialog.getExistingDirectory(self, "Choisir le dossier de sauvegarde")
        if not folder:
            return

        base_name, ok = QInputDialog.getText(
            self,
            "Nom de base",
            "Nom de base du fichier ? (ex: segment => segment_0-99.csv)"
        )

        if not ok or not base_name:
            return

        try:
            for i in range(0, len(data), segment_size):
                start = i
                end = min(i + segment_size, len(data))
                segment = data[start:end]

                file_name = f"{base_name}_{start}-{end - 1}.csv"
                file_path = os.path.join(folder, file_name)

                with open(file_path, mode="w", newline='', encoding='utf-8') as f:
                    writer = csv.writer(f, delimiter=';')
                    writer.writerow(self.headers)
                    writer.writerows(segment)

            self.show_message(f"Segments sauvegardés dans {folder}")

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la sauvegarde : {str(e)}")
